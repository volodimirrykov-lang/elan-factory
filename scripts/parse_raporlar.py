"""
parse_raporlar.py — вытаскивает числа из RAPORLAR xlsx (лист ИТОГ / İCMAL)
и пишет JSON-снапшот в data/snapshots/YYYY-MM.json.

Usage:
    python scripts/parse_raporlar.py _raw/"RAPORLAR 2026 MART Kimya (1).xlsx" 2026-03

Структура ИТОГ листа стабильна месяц к месяцу:
  — Колонки A-E: банковские позиции + итоги (ПОСТУПЛЕНИЯ/ОПЛАТЫ/ПРОДАЖИ/ЭКСПОРТ + клиенты)
  — Колонки G-H: разбивка расходов (сырьё, доставка, фабрика, персонал, услуги, прочее)
Парсер label-driven (ищет по тексту в col A или col G), устойчив к сдвигам строк.
"""
import sys, os, json, re
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

# ── Ярлыки, которые ищем в левой колонке (col A) ──
LEFT_LABELS = {
    # bank positions — col A name, col B start, col C in, col D out, col E end
    "EMLAK BANK TL":            "bank_emlak_tl",
    "VAKIF BANK TL":            "bank_vakif_tl",
    "KASA TL":                  "bank_kasa_tl",
    "EMLAK BANK EURO":          "bank_emlak_eur",
    "VAKIF BANK EURO":          "bank_vakif_eur",
    "KASA EURO":                "bank_kasa_eur",
    "EMLAK BANK USD":           "bank_emlak_usd",
    "VAKIF BANK USD":           "bank_vakif_usd",
    "KASA USD":                 "bank_kasa_usd",
}

# Итоги (только col B нужен)
LEFT_TOTALS = {
    "ПОСТУПЛЕНИЯ TAHSİLATLAR": "cash_in",
    "ОПЛАТЫ ÖDEMELER":         "cash_out",
    "ПРОДАЖИ SATIŞLAR":        "revenue",
    "ЭКСПОРТ":                 "export_total",
}

# Справа (col G) — поступления от учредителей / другое направление (в ELAN KIMYA)
# Суммируются в totals.founder_injection
FOUNDER_PATTERNS = [
    "Поступление от учредителей",
    "Поступление наличными от учредителей",
]

# Выручка по клиентам. Матчится по startswith — порядок важен: спец. варианты первее.
# Блок идёт от "ЭКСПОРТ" до "Дебеторская задолженность". Внутри блока есть под-заголовки
# (ЭКСПОРТ, ТУРЦИЯ, УСЛУГИ) — их пропускаем как subtotal, не матчим.
CLIENT_PATTERNS = [
    ("ФОП Кравченко",          "fop_kravchenko_ua"),
    ("ELAN COSMETİCS GENERAL", "elan_cosmetics_uae"),
    ("ELAN COSMETICS GENERAL", "elan_cosmetics_uae"),
    ("ELAN BEAUTY SPOLKA",     "elan_beauty_pl"),
    ("Австралия",              "australia"),
    ("Приваты",                "privates"),
    ("LUNESİ UK RAYE",         "lunesi_uk_raye"),
    ("LUNESI UK RAYE",         "lunesi_uk_raye"),
    ("LUNESİ COSMOPROF",       "lunesi_cosmoprof"),
    ("LUNESI COSMOPROF",       "lunesi_cosmoprof"),
    ("LUNESİ UK",              "lunesi_uk"),
    ("LUNESI UK",              "lunesi_uk"),
    ("CLAB",                   "clab"),
    ("C-LAB",                  "clab"),
    ("BİLOBROV",               "bilobrov"),
    ("BILOBROV",               "bilobrov"),
    ("USUPSO",                 "usupso"),
    ("EUROPA BELLE",           "europa_belle"),
    ("ELAN KOZMETİK",          "elan_kozmetik_tr"),
    ("ELAN KOZMETIK",          "elan_kozmetik_tr"),
]

# Строки-заголовки подблоков внутри "клиентского" блока — пропускаем, не считаем как клиентов
CLIENT_BLOCK_HEADERS = ("ЭКСПОРТ", "ТУРЦИЯ", "УСЛУГИ (разработка")

# Расходы (col G label → col H amount)
RIGHT_COST_LABELS = {
    "ПРЯМЫЕ ЗАТРАТЫ НА СЫРЬЕ И МАТЕРИАЛЫ": "cogs_materials_total",
    "Закупка материалов":                  "cogs_materials_purchase",
    "Тара":                                "cogs_packaging_jars",
    "Тара из Китая":                       "cogs_packaging_china",
    "Этикетка":                            "cogs_labels",
    "Коробка":                             "cogs_boxes",
    "Закупка сырья":                       "cogs_raw",
    "ЗАТРАТЫ НА ДОСТАВКУ И ОТГРУЗКУ":       "cogs_shipping_total",
    "СОДЕРЖАНИЕ ФАБРИКИ":                  "opex_factory_total",
    "РАСХОДНЫ НА ПЕРСОНАЛ":                "opex_personnel_total",
    "РАСХОДЫ НА ПЕРСОНАЛ":                 "opex_personnel_total",
    "Зарплата":                            "opex_personnel_salary",
    "ЕСВ (SGK)":                           "opex_personnel_sgk",
    "УСЛУГИ":                              "opex_services_total",
    "Тамможенные налоги при импорте":       "customs_duty",
    "Таможенные налоги при импорте":        "customs_duty",
    "ПРОЧИЕ РАСХОДЫ":                      "opex_other_total",
    "Инвестиции оборудование":              "capex_through_opex",
    "Затраты украинского офиса":            "opex_ukr_office",
    "ВСЕГО":                               "grand_total_payments",
}

# Inventory (col A → col B units, col C value)
INVENTORY_PATTERNS = [
    ("Готовая продукция", "finished_goods"),
    ("Сырье",             "raw_materials"),
    ("Тара",              "jars"),
    ("Коробка",           "boxes"),
    ("Этикетка",          "labels"),
    ("Инструкция",        "instructions"),
]


def num(v):
    if v is None: return None
    if isinstance(v, (int, float)): return float(v)
    try:
        return float(str(v).replace(",", ".").replace(" ", ""))
    except ValueError:
        return None


def contains(cell, sub):
    return cell is not None and sub.lower() in str(cell).lower()


def parse(xlsx_path: str, period: str) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    sheet_name = next((s for s in wb.sheetnames if "ИТОГ" in s or "İCMAL" in s), None)
    if not sheet_name:
        raise RuntimeError(f"No ИТОГ/İCMAL sheet in {xlsx_path}")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    out = {
        "period": period,
        "source_file": os.path.basename(xlsx_path),
        "banks": {},
        "totals": {},
        "clients": {},
        "costs": {},
        "inventory": {},
    }

    in_inventory_block = False
    in_clients_block = False   # True только между "ЭКСПОРТ" и "Дебеторская/УСЛУГИ"
    for row in rows:
        # pad row to max 12 cols; первая колонка в xlsx пустая → сдвиг на +1
        row = list(row) + [None] * (12 - len(row)) if len(row) < 12 else list(row)
        a, b, c, d, e = row[1], row[2], row[3], row[4], row[5]
        g, h = row[7], row[8]

        # Bank positions
        if a:
            a_str = str(a).strip()
            for lbl, key in LEFT_LABELS.items():
                if a_str.startswith(lbl):
                    out["banks"][key] = {
                        "label": lbl,
                        "start":  num(b),
                        "in":     num(c),
                        "out":    num(d),
                        "end":    num(e),
                    }
                    break

            # Totals
            for lbl, key in LEFT_TOTALS.items():
                if lbl in a_str:
                    out["totals"][key] = num(b)
                    if lbl == "ЭКСПОРТ":
                        in_clients_block = True
                    break

            # Закрываем блок клиентов на секции Дебеторская (дальше идут receivables)
            if in_clients_block and "Дебеторская" in a_str:
                in_clients_block = False

            # Clients — в блоке между ЭКСПОРТ и Дебеторская
            if in_clients_block:
                # Пропускаем строки-подзаголовки (subtotals)
                if a_str.startswith(CLIENT_BLOCK_HEADERS):
                    pass
                else:
                    matched = False
                    for pat, key in CLIENT_PATTERNS:
                        if a_str.startswith(pat) and num(b) is not None:
                            out["clients"][key] = out["clients"].get(key, 0.0) + num(b)
                            matched = True
                            break
                    # Неопознанные строки в блоке → bucket misc_clients (для валидации)
                    if not matched and num(b) is not None and num(b) > 0:
                        out["clients"]["misc"] = out["clients"].get("misc", 0.0) + num(b)

            # Inventory block (col A: Готовая продукция, Сырье etc)
            if contains(a, "Запасы предприятия"):
                in_inventory_block = True
                continue
            if in_inventory_block:
                for pat, key in INVENTORY_PATTERNS:
                    if pat in a_str:
                        out["inventory"][key] = {"units": num(b), "value": num(c)}
                        break
                if contains(a, "Всего"):
                    in_inventory_block = False

        # Cost categories (col G → col H)
        if g:
            g_str = str(g).strip()
            for lbl, key in RIGHT_COST_LABELS.items():
                if lbl in g_str:
                    out["costs"][key] = num(h)
                    break

            # Founder injections — suммируем
            for pat in FOUNDER_PATTERNS:
                if pat in g_str and num(h) is not None:
                    out["totals"]["founder_injection"] = out["totals"].get("founder_injection", 0.0) + num(h)
                    break

    wb.close()
    return out


EXPECTED_LEFT_LABELS = list(LEFT_LABELS.keys()) + list(LEFT_TOTALS.keys())
EXPECTED_RIGHT_LABELS = list(RIGHT_COST_LABELS.keys())
# Если парсер нашёл < этих процентов ожидаемых ярлыков — xlsx поменял структуру
HEALTH_THRESHOLD = 0.7


def health_check(data: dict):
    """Warning если xlsx поменял структуру (меньше ярлыков матчится)."""
    warnings = []
    banks_found = len(data.get("banks", {}))
    expected_banks = len(LEFT_LABELS)
    if banks_found < expected_banks * HEALTH_THRESHOLD:
        warnings.append(f"⚠️  Найдено только {banks_found}/{expected_banks} банков — формат xlsx мог измениться")

    totals_found = len(data.get("totals", {}))
    expected_totals = len(LEFT_TOTALS)  # cash_in, cash_out, revenue, export_total
    if totals_found < expected_totals * HEALTH_THRESHOLD:
        warnings.append(f"⚠️  Найдено только {totals_found}/{expected_totals} итогов — проверь заголовки")

    costs_found = len(data.get("costs", {}))
    if costs_found < 10:  # минимум ожидаем ~10 категорий расходов
        warnings.append(f"⚠️  Найдено только {costs_found} категорий расходов — проверь col G в xlsx")

    # Sanity: если revenue > 0, должен быть хотя бы один клиент
    revenue = data.get("totals", {}).get("revenue", 0) or 0
    if revenue > 1000 and not data.get("clients"):
        warnings.append(f"⚠️  revenue {revenue:.0f} > 0, но клиентов 0 — блок клиентов не распарсился")

    return warnings


def main():
    if len(sys.argv) < 3:
        print("Usage: python parse_raporlar.py <xlsx_path> <YYYY-MM>")
        sys.exit(1)
    xlsx_path, period = sys.argv[1], sys.argv[2]
    if not os.path.exists(xlsx_path):
        print(f"File not found: {xlsx_path}")
        sys.exit(1)

    try:
        data = parse(xlsx_path, period)
    except Exception as e:
        print(f"❌ Parse error: {e}")
        print(f"   xlsx structure may have changed. Check _raw/ files vs parser patterns.")
        sys.exit(1)

    for w in health_check(data):
        print(w)

    # Write to data/snapshots/YYYY-MM.json
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    out_path = os.path.join(root, "data", "snapshots", f"{period}.json")
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"Wrote {out_path}")
    print(f"  totals:   {data['totals']}")
    print(f"  clients:  {len(data['clients'])} entries")
    print(f"  costs:    {len(data['costs'])} entries")
    print(f"  banks:    {len(data['banks'])} accounts")


if __name__ == "__main__":
    main()
